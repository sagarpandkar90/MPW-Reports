def house_visit_lookup():
    import streamlit as st
    import pandas as pd
    import datetime
    import random
    from pathlib import Path
    from openpyxl import load_workbook

    BASE_DIR        = Path(__file__).resolve().parent
    movement_path   = BASE_DIR / "movement.xlsx"
    mno_path        = BASE_DIR / "Shelgaon_mno_records.xlsx"
    family_path = BASE_DIR / "sample_family_data.xlsx"

    for p in [movement_path, mno_path, family_path ]:
        if not p.exists():
            st.error(f"❌ '{p.name}' project folder मध्ये नाही!")
            return

    @st.cache_data
    def load_movement(path_str):
        wb   = load_workbook(path_str, read_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        month_col_labels = list(header[10:])
        mr_month_map = {
            'जाने': 1,  'फेब': 2,    'मार्च': 3,   'एप्रिल': 4,
            'मे':   5,  'जून': 6,    'जुलै': 7,    'ऑग': 8,
            'सेप्टें': 9, 'ऑक्टो': 10, 'नोव्हे': 11, 'डिसे': 12,
        }
        half_meta = []
        for i, label in enumerate(month_col_labels):
            label = str(label).strip()
            dash  = label.rfind('-')
            mr    = label[:dash].strip() if dash != -1 else label
            code  = int(label[dash+1:].strip()) if dash != -1 else 0
            half_meta.append({
                "label":     label,
                "month_num": mr_month_map.get(mr, 0),
                "part":      1 if i % 2 == 0 else 2,
                "code":      code,
                "col_idx":   i,
            })
        data_rows = []
        for row in rows[1:]:
            gav   = str(row[0]).strip() if row[0] else ""
            vasti = str(row[1]).strip() if row[1] else ""
            if "सुट्टी" in gav:
                continue
            try:
                pasun   = int(row[3])
                paryant = int(row[4])
            except (TypeError, ValueError):
                continue
            date_vals = list(row[10:])
            data_rows.append({
                "vasti":     vasti,
                "pasun":     pasun,
                "paryant":   paryant,
                "date_vals": date_vals,
            })
        return half_meta, data_rows

    @st.cache_data
    def load_family_data(path_str):
        """Returns { mno -> [{name, age, sex, ajar}, ...] }"""
        if not Path(path_str).exists(): return {}
        wb = load_workbook(path_str, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip().lower() if c else "" for c in rows[0]]

        # Identify columns
        idx_mno = next((i for i, h in enumerate(header) if any(x in h for x in ['mno', 'm_no'])), None)
        idx_name = next((i for i, h in enumerate(header) if 'name' in h or 'नाव' in h), None)
        idx_age = next((i for i, h in enumerate(header) if 'age' in h or 'वय' in h or 'birth' in h), None)
        idx_sex = next((i for i, h in enumerate(header) if 'sex' in h or 'ling' in h or 'लिंग' in h), None)
        idx_aj = next((i for i, h in enumerate(header) if 'ajar' in h or 'आजारी' in h), None)

        f_map = {}
        for row in rows[1:]:
            try:
                mno = int(float(row[idx_mno]))
                f_map.setdefault(mno, []).append({
                    "name": str(row[idx_name] or "Unknown"),
                    "age": str(row[idx_age] or "-"),
                    "sex": str(row[idx_sex] or "-"),
                    "ajar": str(row[idx_aj]).strip().lower() if idx_aj else "no"
                })
            except:
                continue
        return f_map

    # ── IMPROVEMENT 1: load total_container from xlsx ──────────
    @st.cache_data
    def load_mno_records(path_str):
        """Returns mno_map: { mno -> (family_head, member_count, total_container) }"""
        wb = load_workbook(path_str, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        try:
            idx_mno  = next(i for i, h in enumerate(header) if 'm_no'         in h)
            idx_fam  = next(i for i, h in enumerate(header) if 'family_head'  in h)
            idx_mem  = next(i for i, h in enumerate(header) if 'member_count' in h)
        except StopIteration:
            st.error("❌ Shelgaon_mno_records.xlsx मध्ये 'm_no', 'family_head' किंवा 'member_count' column सापडले नाही!")
            return {}
        # total_container column — optional, defaults to 0 if absent
        idx_tc = next(
            (i for i, h in enumerate(header) if 'total_container' in h),
            None,
        )
        mno_map = {}
        for row in rows[1:]:
            try:
                mno  = int(row[idx_mno])
                fam  = str(row[idx_fam]).strip() if row[idx_fam] else ""
                mem  = int(row[idx_mem]) if row[idx_mem] is not None else 0
                tc   = int(row[idx_tc]) if (idx_tc is not None and row[idx_tc] is not None) else 0
                mno_map[mno] = (fam, mem, tc)
            except (TypeError, ValueError):
                continue
        return mno_map

    # ─────────────────────────────────────────────────────────────
    # Helpers
    # ─────────────────────────────────────────────────────────────
    def render_family_tables(mno_list, mno_map, family_map):
        st.markdown("### 🏠 आजच्या भेटीतील घरे व कुटुंब सदस्य")
        st.caption("खालील घरांच्या यादीवर क्लिक करून घरातील सर्व सदस्यांची माहिती पहा.")

        for mno in mno_list:
            head = mno_map.get(mno, ("नोंद नाही", 0, 0))[0]
            members = family_map.get(mno, [])

            with st.expander(f"म.क्र. {mno} | {head} | सदस्य: {len(members)}"):
                if not members:
                    st.write("या घराची सदस्य माहिती उपलब्ध नाही.")
                else:
                    rows = ""
                    for m in members:
                        status = "<span style='color:red;'>आजारी</span>" if "yes" in m[
                            'ajar'] else "<span style='color:green;'>ठीक</span>"
                        rows += f"<tr><td>{m['name']}</td><td>{m['sex']}</td><td style='text-align:center;'>{m['age']}</td><td>{status}</td></tr>"

                    st.markdown(f"""
                    <table style="width:100%; border-collapse: collapse; font-family: 'Noto Sans Devanagari';">
                        <tr style="background-color: #f0f2f6;"><th>नाव</th><th>लिंग</th><th>वय</th><th>स्थिती</th></tr>
                        {rows}
                    </table>
                    """, unsafe_allow_html=True)

    def names_to_str(name_list):
        if not name_list:
            return "(नाव उपलब्ध नाही)"
        if len(name_list) == 1:
            return name_list[0]
        return ", ".join(name_list[:-1]) + " आणि " + name_list[-1]

    def get_patient_display(info):
        """Build 'रुग्णाचे नाव (लिंग, वय, इतर)' from info dict."""
        # patient_name takes priority; fallback to kutumb selectbox name
        name = info.get("patient_name", "").strip()
        if not name:
            sel = info.get("name", "")
            if sel not in ("-- कुटुंब प्रमुख निवडा --", ""):
                name = sel
        if not name:
            return ""
        parts = []
        if info.get("gender", "--") not in ("--", ""):
            parts.append(info["gender"])
        if info.get("age", "").strip():
            parts.append(f"वय {info['age'].strip()}")
        if info.get("other", "").strip():
            parts.append(info["other"].strip())
        suffix = f" ({', '.join(parts)})" if parts else ""
        return name + suffix

    # ─────────────────────────────────────────────────────────────
    # Build patient lines — ONLY for conditions that ARE present
    # ─────────────────────────────────────────────────────────────
    def build_patient_lines(
        has_fever,    fever_info,
        has_tb,       tb_info,
        has_dog,      dog_info,
        has_motibi,   motibi_info,
        has_kusthar,  kusthar_info,
        has_motibi_op, motibi_op_info,
    ):
        lines = []

        if has_fever:
            fever_name = get_patient_display(fever_info)
            if fever_name:
                lines.append(random.choice([
                    f"{fever_name} यांना ताप असल्याचे आढळले. त्यांना रक्त नमुना तपासणीसाठी प्राथमिक आरोग्य केंद्रात पाठविण्यात आले.",
                    f"{fever_name} यांना तापाने आजारी असल्याचे निदर्शनास आले. हिवताप स्लाइड तयार करून त्यांना वैद्यकीय तपासणीस पाठविले.",
                    f"{fever_name} हे संशयित हिवताप / डेंग्यू रुग्ण आढळले. रुग्णाची माहिती वरिष्ठ कार्यालयास कळविण्यात आली व उपचार सुरू केले.",
                    f"{fever_name} यांना थंडी व ताप असल्याचे समजले. त्यांचा रक्त नमुना घेण्यात आला.",
                ]))

        if has_tb:
            tb_name = get_patient_display(tb_info)
            if tb_name:
                lines.append(random.choice([
                    f"{tb_name} यांना दीर्घकाळ खोकला व वजन कमी होत असल्याची तक्रार असल्याने त्यांना संशयित क्षयरोग (TB) म्हणून नोंद करून DOTS केंद्रात पाठविण्यात आले.",
                    f"{tb_name} यांच्यात क्षयरोगाची संशयित लक्षणे आढळली. त्यांना थुंकी तपासणीसाठी आरोग्य केंद्रात संदर्भित केले.",
                    f"{tb_name} यांना दोन आठवड्यांपेक्षा जास्त काळ खोकला असल्याचे समजले. TB संशयित म्हणून नोंद करून आवश्यक तपासणी केली.",
                ]))

        if has_dog:
            dog_name = get_patient_display(dog_info)
            if dog_name:
                lines.append(random.choice([
                    f"{dog_name} यांना कुत्रा चावल्याची घटना निदर्शनास आली. त्यांना तातडीने अँटी-रेबीज लस घेण्यासाठी प्राथमिक आरोग्य केंद्रात पाठविण्यात आले.",
                    f"{dog_name} हे कुत्रा चावल्यामुळे जखमी झाल्याचे आढळले. ARV उपचारासाठी रुग्णालयात पाठविले.",
                    f"{dog_name} हे संशयित श्वानदंश रुग्ण आढळले. त्वरित अँटी-रेबीज उपचारासाठी संदर्भित करून वरिष्ठांना माहिती दिली.",
                ]))

        if has_motibi:
            motibi_name = get_patient_display(motibi_info)
            if motibi_name:
                lines.append(random.choice([
                    f"{motibi_name} यांना दृष्टी कमी होत असल्याचे व डोळ्यांत पांढरा पडदा येत असल्याचे आढळले. संशयित मोतीबिंदू म्हणून नोंद करून नेत्र तपासणीसाठी जिल्हा रुग्णालयात पाठविण्यात आले.",
                    f"{motibi_name} यांना वाचण्यास व जवळचे पाहण्यास त्रास होत असल्याचे आढळले. नेत्रतज्ज्ञांकडे तपासणीसाठी संदर्भित केले.",
                    f"{motibi_name} यांच्या डोळ्यांवर संशयित मोतीबिंदूची लक्षणे दिसल्याने नोंद घेऊन योग्य उपचारासाठी मार्गदर्शन केले.",
                ]))

        if has_kusthar:
            kusthar_name = get_patient_display(kusthar_info)
            if kusthar_name:
                lines.append(random.choice([
                    f"{kusthar_name} यांच्या त्वचेवर बधिर चट्टे व संवेदनाहीन भाग असल्याचे निदर्शनास आले. संशयित कुष्ठरोग म्हणून नोंद करून NLEP केंद्रात तपासणीसाठी पाठविण्यात आले.",
                    f"{kusthar_name} यांच्या हाता-पायावर बधिरपणा व फिकट चट्टे आढळले. कुष्ठरोग संशयित म्हणून नोंद घेऊन जिल्हा कुष्ठरोग अधिकाऱ्यांना माहिती दिली.",
                    f"{kusthar_name} यांना त्वचेवरील चट्ट्यांबद्दल विचारले असता बधिरपणाची तक्रार असल्याचे कळले. संशयित कुष्ठरोग म्हणून नोंद करून उपचारासाठी संदर्भित केले.",
                ]))

        if has_motibi_op:
            motibi_op_name = get_patient_display(motibi_op_info)
            if motibi_op_name:
                lines.append(random.choice([
                    f"{motibi_op_name} यांनी यापूर्वी मोतीबिंदू शस्त्रक्रिया केल्याचे सांगितले. शस्त्रक्रियेनंतरची प्रकृती तपासली; ते बरे असल्याचे समजले व नोंद घेण्यात आली.",
                    f"{motibi_op_name} यांची मोतीबिंदू ऑपरेशन नंतरची प्रकृती उत्तम असल्याचे आढळले. नियमित डोळ्याचे थेंब वापरण्याचे व पुनर्भेटीसाठी येण्याचे निर्देश दिले.",
                    f"{motibi_op_name} यांनी यशस्वी मोतीबिंदू शस्त्रक्रिया झाल्याचे सांगितले. दृष्टी सुधारल्याचे त्यांनी सांगितले; माहिती नोंदविण्यात आली.",
                ]))

        return "\n".join(lines)

    # ─────────────────────────────────────────────────────────────
    # 10 Diary Formats — unchanged from original
    # ─────────────────────────────────────────────────────────────
    def get_diary_formats():
        return [
            # Format 1
            lambda d: (
                f"📅 दिनांक: {d['date']}\n"
                f"📍 वस्ती: {d['vasti']}\n\n"
                f"आज दिनांक {d['date']} रोजी मी {d['vasti']} येथे घरभेट सर्वेक्षण केले. "
                f"सदर भेटीत म.क्र. {d['pasun']} ते {d['paryant']} पैकी {d['visited_count']} घरांना भेट दिली. "
                f"सकाळी वेळेवर निघून प्रत्येक घरी जाऊन कुटुंब प्रमुखांशी संवाद साधला व त्यांच्या घरातील पाणी साठ्यांची तपासणी केली.\n\n"
                f"आज {d['names_str']} यांच्या घरांना भेट देऊन सर्वेक्षण पूर्ण केले. "
                f"सर्वांनी अत्यंत सहकार्य केले व विचारलेल्या प्रश्नांना उत्तरे दिली. "
                f"प्रत्येक कुटुंबाला आठवड्यातून एकदा पाण्याची भांडी रिकामी करण्याचे व घर-परिसर स्वच्छ ठेवण्याचे महत्त्व पटवून दिले. "
                f"डास उत्पत्ती रोखण्यासाठी साचलेले पाणी काढून टाकण्याच्या सवयीबाबत जनजागृती केली.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"एकूण {d['total_containers']} कंटेनर तपासले, त्यापैकी {d['das_containers']} कंटेनरमध्ये डासांच्या अळ्या आढळून आल्या "
                f"व {d['rikame_containers']} कंटेनर रिकामे करण्यात आले. "
                f"पॉझिटिव्ह आढळलेल्या भांड्यांमधील पाणी ओतून टाकण्यात आले व संबंधितांना पुन्हा असे होणार नाही याची खबरदारी घेण्यास सांगितले.\n\n"
                f"House Index: {d['house_index']}% | Container Index: {d['container_index']}% | Breteau Index: {d['breteau_index']}%"
            ),
            # Format 2
            lambda d: (
                f"📅 दिनांक: {d['date']}\n"
                f"📍 ठिकाण: {d['vasti']}\n\n"
                f"आजच्या दिवशी {d['vasti']} वस्तीतील नियोजित घरभेट कार्यक्रम पार पडला. "
                f"म.क्र. {d['pasun']} ते {d['paryant']} या क्रमांकातील {d['visited_count']} घरांचे सर्वेक्षण करण्यात आले. "
                f"सर्वेक्षणाच्या सुरुवातीला वस्तीतील रस्त्यावरून फेरफटका मारून परिसराचे निरीक्षण केले. "
                f"ठिकठिकाणी साचलेले पाणी व टाकाऊ वस्तू आढळल्या; त्याबाबत संबंधितांना सूचना दिल्या.\n\n"
                f"भेट दिलेल्या प्रमुख कुटुंबांमध्ये {d['names_str']} यांचा समावेश होता. "
                f"सर्व घरांतून माहिती संकलन करण्यात आले. "
                f"कुटुंब प्रमुखांशी हिवताप, डेंग्यू व चिकुनगुनियाच्या प्रतिबंधाबाबत सविस्तर चर्चा करण्यात आली. "
                f"डासांची पैदास रोखण्यासाठी घरात व अवतीभवती पाणी साचू देऊ नये असे बजावण्यात आले.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"आज एकूण {d['total_containers']} पाणी साठवण्याच्या भांड्यांची तपासणी झाली. "
                f"{d['das_containers']} ठिकाणी डासांची उत्पत्तीस्थाने आढळली आणि {d['rikame_containers']} भांडी रिकामी करण्यात आली. "
                f"उर्वरित भांडी स्वच्छ असल्याचे आढळले.\n\n"
                f"निर्देशांक — HI: {d['house_index']}% | CI: {d['container_index']}% | BI: {d['breteau_index']}%"
            ),
            # Format 3
            lambda d: (
                f"🗓️ दिनांक {d['date']} — क्षेत्र भेट नोंद\n"
                f"📌 वस्ती: {d['vasti']}\n\n"
                f"आज {d['vasti']} येथे किटकशास्त्रीय सर्वेक्षणासाठी भेट दिली. "
                f"म.क्र. {d['pasun']} पासून {d['paryant']} पर्यंत {d['visited_count']} घरांचे निरीक्षण केले. "
                f"{d['names_str']} या कुटुंब प्रमुखांशी भेट होऊन त्यांना डास प्रतिबंधाबाबत मार्गदर्शन केले. "
                f"वस्तीतील काही घरांमागे मोकळ्या जागी टाकाऊ टायर व डबे पडलेले आढळले; ते हटविण्याची विनंती केली.\n\n"
                f"सर्वेक्षणादरम्यान रहिवाशांनी चांगले सहकार्य दिले. "
                f"ज्या घरांत कोणी नव्हते त्यांना नंतर भेट देण्याचे नियोजन केले. "
                f"प्रत्येक घराचे कंटेनर तपासताना माठ, बादल्या, टाक्या, कुलर व फुलदाण्या यांचाही समावेश केला.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"तपासणी सारांश: {d['total_containers']} कंटेनर तपासले → "
                f"{d['das_containers']} मध्ये अळ्या आढळल्या → {d['rikame_containers']} रिकामे केले. "
                f"सकारात्मक घरांच्या मालकांना लेखी सूचना देण्यात आली.\n\n"
                f"HI: {d['house_index']}% | CI: {d['container_index']}% | BI: {d['breteau_index']}%"
            ),
            # Format 4
            lambda d: (
                f"दैनंदिन नोंद — {d['date']}\n"
                f"स्थळ: {d['vasti']} | म.क्र. श्रेणी: {d['pasun']} – {d['paryant']}\n\n"
                f"आज सकाळी {d['vasti']} वस्तीत घरोघरी जाऊन सर्वेक्षण केले. "
                f"{d['names_str']} इत्यादी कुटुंबांना भेटून त्यांच्या घरातील व परिसरातील पाणी साठ्यांची तपासणी केली. "
                f"रहिवाशांनी उत्तम सहकार्य दिले. काही ठिकाणी घरमालक स्वतः पुढे येऊन भांडी दाखवत होते, "
                f"हे पाहून जनजागृतीचा सकारात्मक परिणाम जाणवला.\n\n"
                f"वस्तीत एकूण {d['visited_count']} घरांना भेट दिली. "
                f"प्रत्येक घरात किमान पाच ते सहा कंटेनर तपासले. "
                f"काही घरांत पाण्याचे हौद उघडे असल्याचे आढळले; ते झाकण्याची सूचना दिली. "
                f"घराबाहेरील टाकाऊ डबे व जुने टायर हटविण्यास सांगितले.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"कंटेनर तपासणी: एकूण {d['total_containers']} | डास आढळले {d['das_containers']} | रिकामे केले {d['rikame_containers']}\n\n"
                f"House Index: {d['house_index']}% | Container Index: {d['container_index']}% | Breteau Index: {d['breteau_index']}%"
            ),
            # Format 5
            lambda d: (
                f"📋 घरभेट दैनंदिनी\n"
                f"📅 दिनांक: {d['date']} | 📍 वस्ती: {d['vasti']}\n\n"
                f"आज {d['vasti']} येथील म.क्र. {d['pasun']} ते {d['paryant']} पर्यंतच्या {d['visited_count']} घरांना भेट दिली. "
                f"वस्तीत पोहोचल्यावर प्रथम वस्ती प्रमुखांशी भेट घेऊन सर्वेक्षणाची माहिती दिली. "
                f"त्यांनी सकारात्मक प्रतिसाद दिला व रहिवाशांना सहकार्य करण्यास सांगितले.\n\n"
                f"भेट घेतलेले कुटुंब प्रमुख: {d['names_str']}\n\n"
                f"या सर्व कुटुंबांना हिवताप व डेंग्यू प्रतिबंधाबाबत जनजागृती केली. "
                f"घरातील व अवतीभवतीच्या कंटेनरची तपासणी केली. "
                f"डासांच्या जीवनचक्राबद्दल सोप्या भाषेत माहिती सांगितली व घरोघरी लीफलेट वाटप केले. "
                f"उघड्या विहिरींवर व नाल्यांजवळील परिस्थितीची नोंद घेतली.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"तपासणी निकाल — तपासलेले: {d['total_containers']} | "
                f"डास आढळलेले: {d['das_containers']} | "
                f"रिकामे केलेले: {d['rikame_containers']}\n\n"
                f"निर्देशांक: HI = {d['house_index']}% | CI = {d['container_index']}% | BI = {d['breteau_index']}%"
            ),
            # Format 6
            lambda d: (
                f"दिनांक: {d['date']}\n"
                f"वस्ती: {d['vasti']}\n\n"
                f"आज {d['vasti']} येथे नियमित घरभेट सर्वेक्षण करण्यात आले. "
                f"म.क्र. {d['pasun']} ते {d['paryant']} या क्रमांकाच्या {d['visited_count']} घरांना भेट देऊन किटकशास्त्रीय तपासणी केली. "
                f"वस्तीत पोहोचताना उन्हाचा जोर होता; तरीही रहिवाशांनी सहकार्य करून घरे उघडून दिली.\n\n"
                f"{d['names_str']} यांच्या घरांची तपासणी करताना पाण्याची भांडी, टाक्या व इतर साठवण स्थळांची "
                f"बारकाईने पाहणी केली. रहिवाशांना डास उत्पत्ती रोखण्यासाठी उपाययोजनांची माहिती दिली. "
                f"काही घरांत फ्रिजच्या ट्रे व एसी च्या पाण्याची ट्रे यांमध्ये पाणी साचलेले आढळले; "
                f"त्याबाबत विशेष लक्ष देण्यास सांगितले. फुलदाण्यांमधील पाणी दर दोन दिवसांनी बदलण्याची सूचना दिली.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"सर्वेक्षण निष्कर्ष: {d['total_containers']} कंटेनर तपासले, "
                f"{d['das_containers']} मध्ये डासांच्या अळ्या आढळल्या, "
                f"{d['rikame_containers']} कंटेनर रिकामे व स्वच्छ करण्यात आले.\n\n"
                f"House Index: {d['house_index']}% | Container Index: {d['container_index']}% | Breteau Index: {d['breteau_index']}%"
            ),
            # Format 7
            lambda d: (
                f"🗒️ क्षेत्र सर्वेक्षण नोंद | {d['date']}\n\n"
                f"आजचे सर्वेक्षण क्षेत्र: {d['vasti']} (म.क्र. {d['pasun']} – {d['paryant']})\n\n"
                f"{d['vasti']} वस्तीत आज {d['visited_count']} घरांना भेट देण्यात आली. "
                f"{d['names_str']} या कुटुंब प्रमुखांसमवेत चर्चा करून त्यांना हिवताप व डेंग्यू रोगाच्या "
                f"प्रतिबंधासाठी घराभोवतालचे पाणी साचणार नाही याची काळजी घेण्यास सांगण्यात आले. "
                f"प्रत्येक कुटुंबाला आठवड्यातून एकदा सर्व पाण्याची भांडी रिकामी करून पुसून ठेवण्याचे महत्त्व पटवून दिले.\n\n"
                f"वस्तीतील काही घरांमागे मोकळ्या जागेत पाण्याचे डबके आढळले. "
                f"ग्रामपंचायतीशी संपर्क साधून निचरा व्यवस्था सुधारण्याची विनंती करण्याचे नियोजन केले.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"आजच्या तपासणीत {d['total_containers']} कंटेनर तपासले. "
                f"{d['das_containers']} मध्ये डासांची उत्पत्ती आढळली व {d['rikame_containers']} तातडीने रिकामे केले. "
                f"पुढील भेटीत याच घरांचा पाठपुरावा करण्याचे नियोजन केले.\n\n"
                f"HI: {d['house_index']}% | CI: {d['container_index']}% | BI: {d['breteau_index']}%"
            ),
            # Format 8
            lambda d: (
                f"{d['date']} — दैनंदिन कार्य नोंद\n"
                f"वस्ती: {d['vasti']} | म.क्र.: {d['pasun']} ते {d['paryant']}\n\n"
                f"आजच्या घरभेटीत {d['vasti']} वस्तीतील {d['visited_count']} घरांचे सर्वेक्षण करण्यात आले. "
                f"{d['names_str']} यांच्याशी संपर्क साधून कुटुंबातील आरोग्य स्थितीची माहिती घेतली "
                f"तसेच डास प्रतिबंधक उपायांबाबत प्रबोधन केले. "
                f"सर्वेक्षणापूर्वी गावाच्या आरोग्य सेविकेशी संपर्क साधून वस्तीतील अलीकडच्या आजारपणाची माहिती घेतली.\n\n"
                f"भेटीत लक्षात आले की काही घरांत पाण्याच्या टाक्या वर्षानुवर्षे साफ केलेल्या नाहीत. "
                f"त्यांना टाकी साफ करण्याचे व झाकण ठेवण्याचे निर्देश दिले.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"भांडी तपासणी अहवाल —\n"
                f"▸ एकूण तपासलेले कंटेनर: {d['total_containers']}\n"
                f"▸ डास आढळलेले कंटेनर: {d['das_containers']}\n"
                f"▸ रिकामे केलेले कंटेनर: {d['rikame_containers']}\n\n"
                f"House Index: {d['house_index']}% | Container Index: {d['container_index']}% | Breteau Index: {d['breteau_index']}%"
            ),
            # Format 9
            lambda d: (
                f"📌 {d['vasti']} — घरभेट | {d['date']}\n\n"
                f"आज {d['date']} रोजी {d['vasti']} वस्तीतील नियोजित घरभेट पार पडली. "
                f"म.क्र. {d['pasun']} ते {d['paryant']} या घरांचे सर्वेक्षण करण्यात आले. "
                f"वस्तीत एकूण {d['visited_count']} घरे असून प्रत्येक घराला किमान पंधरा मिनिटे देऊन "
                f"सविस्तर तपासणी व समुपदेशन केले.\n\n"
                f"आजच्या भेटीत {d['names_str']} या कुटुंबांशी संपर्क झाला. "
                f"पाण्याचे साठे, टाक्या, माठ, बादल्या यांची सखोल तपासणी केली. "
                f"घराच्या आसपास उगवलेल्या झुडपांत व रिकाम्या भांड्यांत पावसाचे पाणी साचलेले आढळले; "
                f"ते त्वरित काढून टाकण्यास सांगितले.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"कंटेनर: {d['total_containers']} तपासले → {d['das_containers']} पॉझिटिव्ह → {d['rikame_containers']} रिकामे. "
                f"पुढील पंधरवड्यात याच घरांची पुन्हा तपासणी करण्याचे नियोजन केले.\n\n"
                f"HI {d['house_index']}% | CI {d['container_index']}% | BI {d['breteau_index']}%"
            ),
            # Format 10
            lambda d: (
                f"🏥 हिवताप / डेंग्यू सर्वेक्षण डायरी\n"
                f"📅 {d['date']} | 📍 {d['vasti']}\n\n"
                f"आज {d['vasti']} येथे किटकशास्त्रीय घरभेट सर्वेक्षण करण्यात आले. "
                f"म.क्र. {d['pasun']} पासून {d['paryant']} पर्यंत एकूण {d['visited_count']} घरांना भेट दिली. "
                f"सर्वेक्षण सुरू करण्यापूर्वी वस्तीतील काही जुन्या रहिवाशांशी बोलून मागील हंगामातील "
                f"ताप व हिवतापाच्या रुग्णांबाबत माहिती संकलित केली.\n\n"
                f"{d['names_str']} या कुटुंब प्रमुखांशी भेट होऊन त्यांच्या घरातील व अंगणातील "
                f"पाणीसाठ्यांची तपासणी केली. प्रत्येक कुटुंबाला आठवड्यातून एकदा पाण्याची भांडी "
                f"रिकामी करण्याचा सल्ला देण्यात आला.\n\n"
                + (f"{d['patient_text']}\n\n" if d['patient_text'] else "")
                + f"कंटेनर तपासणी सारांश:\n"
                f"तपासलेले: {d['total_containers']} | डास आढळलेले: {d['das_containers']} | रिकामे केलेले: {d['rikame_containers']}\n\n"
                f"House Index: {d['house_index']}% | Container Index: {d['container_index']}% | Breteau Index: {d['breteau_index']}%"
            ),
        ]

    # ══════════════════════════════════════════════════════════════
    # MAIN LOGIC
    # ══════════════════════════════════════════════════════════════
    half_meta, data_rows = load_movement(str(movement_path))
    mno_map              = load_mno_records(str(mno_path))
    family_map = load_family_data(str(family_path))
    if not mno_map:
        return

    st.subheader("📋 घरभेट यादी – तारीख निवडा")

    selected_date = st.date_input(
        "तारीख निवडा",
        value=datetime.date.today(),
        min_value=datetime.date(2026, 1, 1),
        max_value=datetime.date(2026, 12, 31),
    )

    if not selected_date:
        return

    sel_month = selected_date.month
    sel_day   = selected_date.day
    sel_part  = 1 if sel_day <= 15 else 2

    matching_half = next(
        (h for h in half_meta if h["month_num"] == sel_month and h["part"] == sel_part),
        None
    )
    if not matching_half:
        st.warning("या तारखेसाठी movement.xlsx मध्ये डेटा नाही.")
        return

    col_idx    = matching_half["col_idx"]
    month_code = matching_half["code"]

    matched_row = None
    for row in data_rows:
        if col_idx < len(row["date_vals"]):
            val = row["date_vals"][col_idx]
            try:
                if int(val) == sel_day:
                    matched_row = row
                    break
            except (TypeError, ValueError):
                continue

    if not matched_row:
        st.info(
            f"📅 **{selected_date.strftime('%d-%m-%Y')}** – या तारखेला कोणत्याही वस्तीची भेट नाही.\n"
            f"(Movement sheet मध्ये या दिवशी visit date नोंदलेली नाही)"
        )
        return

    vasti   = matched_row["vasti"]
    pasun   = matched_row["pasun"]
    paryant = matched_row["paryant"]


    # mno_list generation
    if month_code == 0:
        first = ((pasun + 9) // 10) * 10
    else:
        remainder = pasun % 10
        diff      = (month_code - remainder) % 10
        first     = pasun + diff

    mno_list = list(range(first, paryant + 1, 10))

    # ── IMPROVEMENT 1: pre-fill total_container from xlsx ──────
    results = []
    for idx, mno in enumerate(mno_list, start=1):
        record = mno_map.get(mno, None)
        fam    = record[0] if record else "❓ नोंद नाही"
        mem    = record[1] if record else 0          # always int
        tc     = record[2] if record else 0          # total_container from xlsx
        results.append({
            "अ.क्र.":                              idx,
            "म.क्र. (m_no)":                       mno,
            "कुटुंब प्रमुख":                        fam,
            "सदस्य संख्या":                         mem,
            "तपासलेल्या कंटेनरची संख्या":           tc,  # pre-filled from xlsx
            "डास आढळून आलेल्या कंटेनरची संख्या":    0,
            "रिकामे केलेल्या कंटेनरची संख्या":       0,
        })

    st.markdown(f"""
---
### 📍 {selected_date.strftime('%d-%m-%Y')} – **{vasti}**
| | |
|---|---|
| **Half column** | `{matching_half['label']}` (code: **{month_code}**) |
| **पासून** | {pasun} |
| **पर्यंत** | {paryant} |
| **एकूण घरे (पूर्ण श्रेणी)** | {paryant - pasun + 1} |
""")

    st.markdown("### 👨‍👩‍👧 दैनिक किटकशास्त्रीय सर्वेक्षण")
    st.caption("'तपासलेल्या कंटेनरची संख्या' Shelgaon_mno_records मधील total_container वरून आपोआप भरली आहे — आवश्यक असल्यास बदला.")

    edited_df = st.data_editor(
        pd.DataFrame(results),
        use_container_width=True,
        hide_index=True,
        column_config={
            "अ.क्र.":                              st.column_config.NumberColumn(disabled=True),
            "म.क्र. (m_no)":                       st.column_config.NumberColumn(disabled=True),
            "कुटुंब प्रमुख":                        st.column_config.TextColumn(disabled=True),
            "सदस्य संख्या":                         st.column_config.NumberColumn(disabled=True),
            "तपासलेल्या कंटेनरची संख्या":           st.column_config.NumberColumn(min_value=0, step=1, required=True),
            "डास आढळून आलेल्या कंटेनरची संख्या":    st.column_config.NumberColumn(min_value=0, step=1, required=True),
            "रिकामे केलेल्या कंटेनरची संख्या":       st.column_config.NumberColumn(min_value=0, step=1, required=True),
        },
        key=f"survey_table_{selected_date}",
    )

    # ── IMPROVEMENT 2: editable summary table for indexes ──────
    total_range_ghare = paryant - pasun + 1

    # Auto-compute from per-house table
    auto_das_ghare   = int((edited_df["डास आढळून आलेल्या कंटेनरची संख्या"] > 0).sum())
    auto_total_cont  = int(edited_df["तपासलेल्या कंटेनरची संख्या"].sum())
    auto_das_cont    = int(edited_df["डास आढळून आलेल्या कंटेनरची संख्या"].sum())
    auto_rikame_cont = int(edited_df["रिकामे केलेल्या कंटेनरची संख्या"].sum())

    st.markdown("---")
    st.markdown("### 📊 किटकशास्त्रीय निर्देशांक — एकूण सारांश")
    st.caption("खालील आकडे घरनिहाय तक्त्यावरून आपोआप भरले आहेत. आवश्यक असल्यास थेट बदला.")

    summary_df = pd.DataFrame([{
        "एकूण घरे":                  total_range_ghare,
        "डास आढळलेली घरे":           auto_das_ghare,
        "एकूण तपासलेले कंटेनर":      auto_total_cont,
        "एकूण डास आढळलेले कंटेनर":   auto_das_cont,
        "एकूण रिकामे केलेले कंटेनर": auto_rikame_cont,
    }])

    edited_summary = st.data_editor(
        summary_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            col: st.column_config.NumberColumn(min_value=0, step=1)
            for col in summary_df.columns
        },
        key=f"summary_table_{selected_date}",
    )

    # Compute indexes from editable summary row
    s                     = edited_summary.iloc[0]
    final_ghare           = int(s["एकूण घरे"])
    final_das_ghare       = int(s["डास आढळलेली घरे"])
    final_total_cont      = int(s["एकूण तपासलेले कंटेनर"])
    final_das_cont        = int(s["एकूण डास आढळलेले कंटेनर"])
    final_rikame_cont     = int(s["एकूण रिकामे केलेले कंटेनर"])

    house_index     = (final_das_ghare  / final_ghare      * 100) if final_ghare      > 0 else 0.0
    container_index = (final_das_cont   / final_total_cont * 100) if final_total_cont > 0 else 0.0
    breteau_index   = (final_das_cont   / final_ghare      * 100) if final_ghare      > 0 else 0.0

    c1, c2, c3 = st.columns(3)
    with c1:
        st.metric("🏠 House Index",     f"{house_index:.2f}%",
                  help="(डास आढळलेली घरे ÷ एकूण घरे) × 100")
    with c2:
        st.metric("🪣 Container Index", f"{container_index:.2f}%",
                  help="(डास आढळलेले कंटेनर ÷ एकूण तपासलेले कंटेनर) × 100")
    with c3:
        st.metric("📐 Breteau Index",   f"{breteau_index:.2f}%",
                  help="(डास आढळलेले कंटेनर ÷ एकूण घरे) × 100")

    # ══════════════════════════════════════════════════════════════
    # DIARY SECTION
    # ══════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("### 📝 डायरी लेखन — रुग्ण माहिती")

    # ── IMPROVEMENT 3: names from mno_list range only ──────────
    # Selectbox label = "m_no – कुटुंब प्रमुख" so user can identify house
    mno_label_opts = []
    mno_list_all = list(range(pasun, paryant + 1))
    for m in mno_list_all:
        rec = mno_map.get(m)
        if rec and rec[0] and "नोंद नाही" not in rec[0]:
            mno_label_opts.append(f"{m} – {rec[0]}")

    kutumb_opts = ["-- निवडा --"] + mno_label_opts

    # ── Patient widget ──────────────────────────────────────────
    def rugna_widget(checkbox_label, key_prefix):
        found = st.checkbox(
            checkbox_label,
            value=False,
            key=f"{key_prefix}_chk_{selected_date}",
        )
        info = {
            "name":         "",   # kutumb pramukh (house identifier)
            "patient_name": "",   # actual patient name (free text)
            "gender":       "--",
            "age":          "",
            "other":        "",
        }
        if found:
            with st.container():
                # Row 1: kutumb selectbox + patient name text input
                col_kp, col_pn = st.columns(2)
                with col_kp:
                    sel = st.selectbox(
                        "🏠 कुटुंब प्रमुख निवडा (घर ओळखण्यासाठी):",
                        options=kutumb_opts,
                        key=f"{key_prefix}_kp_{selected_date}",
                    )
                    # Store only the name portion after " – "
                    if sel != "-- निवडा --" and " – " in sel:
                        info["name"] = sel.split(" – ", 1)[1].strip()
                    elif sel != "-- निवडा --":
                        info["name"] = sel
                with col_pn:
                    info["patient_name"] = st.text_input(
                        "👤 रुग्णाचे नाव (स्वतः टाका):",
                        key=f"{key_prefix}_pn_{selected_date}",
                        placeholder="रुग्णाचे पूर्ण नाव लिहा",
                    )
                # Row 2: gender / age / other
                col_g, col_a, col_o = st.columns(3)
                with col_g:
                    info["gender"] = st.selectbox(
                        "लिंग:",
                        options=["--", "पुरुष", "स्त्री", "इतर"],
                        key=f"{key_prefix}_gen_{selected_date}",
                    )
                with col_a:
                    info["age"] = st.text_input(
                        "वय (वर्षे):",
                        key=f"{key_prefix}_age_{selected_date}",
                        placeholder="उदा. 35",
                    )
                with col_o:
                    info["other"] = st.text_input(
                        "इतर माहिती / लक्षणे:",
                        key=f"{key_prefix}_oth_{selected_date}",
                        placeholder="उदा. थंडी ताप 3 दिवस",
                    )
        return found, info

    col_a, col_b = st.columns(2)
    with col_a:
        has_fever,     fever_info     = rugna_widget("🤒 संशयित हिवताप / डेंग्यू रुग्ण आढळला का?",        "fever")
        has_tb,        tb_info        = rugna_widget("🫁 संशयित क्षयरोग (TB) रुग्ण आढळला का?",            "tb")
        has_dog,       dog_info       = rugna_widget("🐕 संशयित श्वानदंश (कुत्रा चावणे) रुग्ण आढळला का?", "dog")
    with col_b:
        has_motibi,    motibi_info    = rugna_widget("👁️ संशयित मोतीबिंदू रुग्ण आढळला का?",              "motibi")
        has_kusthar,   kusthar_info   = rugna_widget("🩹 संशयित कुष्ठरोग रुग्ण आढळला का?",               "kusthar")
        has_motibi_op, motibi_op_info = rugna_widget("🏥 मोतीबिंदू ऑपरेशन केलेला रुग्ण आढळला का?",       "motibi_op")

    st.markdown("")
    if st.button("📔 डायरी लिहा", type="primary", key=f"diary_btn_{selected_date}"):

        # 4-5 random kutumb pramukh names from range
        pool = [
            mno_map[m][0]
            for m in mno_list
            if m in mno_map and mno_map[m][0] and "नोंद नाही" not in mno_map[m][0]
        ]
        if not pool:
            pool = ["(नाव उपलब्ध नाही)"]
        pick_count   = min(random.randint(4, 5), len(pool))
        picked_names = random.sample(pool, pick_count)
        names_str    = names_to_str(picked_names)

        patient_text = build_patient_lines(
            has_fever,    fever_info,
            has_tb,       tb_info,
            has_dog,      dog_info,
            has_motibi,   motibi_info,
            has_kusthar,  kusthar_info,
            has_motibi_op, motibi_op_info,
        )

        diary_data = {
            "date":              selected_date.strftime("%d-%m-%Y"),
            "vasti":             vasti,
            "pasun":             pasun,
            "paryant":           paryant,
            "visited_count":     final_ghare,
            "names_str":         names_str,
            "total_containers":  final_total_cont,
            "das_containers":    final_das_cont,
            "rikame_containers": final_rikame_cont,
            "house_index":       round(house_index, 2),
            "container_index":   round(container_index, 2),
            "breteau_index":     round(breteau_index, 2),
            "patient_text":      patient_text,
        }

        diary_text = random.choice(get_diary_formats())(diary_data)

        st.markdown("### 📔 आजची डायरी")
        safe_text = diary_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        st.markdown(
            "<div style='"
            "background-color:#fffde7;"
            "border-left:5px solid #f9a825;"
            "border-radius:8px;"
            "padding:20px 24px;"
            "font-family:Noto Sans Devanagari,sans-serif;"
            "font-size:15px;"
            "line-height:2.0;"
            "color:#333;"
            "white-space:pre-wrap;"
            "'>"
            + safe_text
            + "</div>",
            unsafe_allow_html=True,
        )

        with st.expander("📋 डायरी कॉपी करा (plain text)"):
            st.text_area(
                label="",
                value=diary_text.strip(),
                height=320,
                key=f"diary_copy_{selected_date}_{random.randint(1000, 9999)}",
            )
    render_family_tables(list(range(pasun, paryant + 1)), mno_map, family_map)
