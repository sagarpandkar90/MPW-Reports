def monthly_diary():
    import streamlit as st
    import pandas as pd
    import calendar
    import datetime
    from pathlib import Path
    import base64
    import json
    import streamlit.components.v1 as components
    from openpyxl import load_workbook

    st.title("मासिक डायरी – Arogya Sevak")

    BASE_DIR      = Path(__file__).resolve().parent
    excel_path    = BASE_DIR / "movement.xlsx"
    holiday_path  = BASE_DIR / "Holidays.xlsx"

    if not excel_path.exists():
        st.error("❌ 'movement.xlsx' project folder मध्ये नाही!")
        return

    # ── Load movement.xlsx (cached) ───────────────────────────────────────────
    @st.cache_data
    def load_movement(path_str):
        wb   = load_workbook(path_str, read_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header     = rows[0]
        month_cols = list(header[10:])

        mr_month_map = {
            'जाने': 1, 'फेब': 2,    'मार्च': 3,  'एप्रिल': 4,
            'मे':   5, 'जून': 6,    'जुलै': 7,   'ऑग': 8,
            'सेप्टें': 9, 'ऑक्टो': 10, 'नोव्हे': 11, 'डिसे': 12,
        }

        def parse_holiday_cell(val):
            if val is None:
                return set()
            result = set()
            for part in str(val).replace(" ", "").split(","):
                if part.isdigit():
                    result.add(int(part))
            return result

        data_rows   = []
        holiday_row = None
        for row in rows[1:]:
            gav   = str(row[0]).strip() if row[0] else ""
            vasti = str(row[1]).strip() if row[1] else ""
            if "सुट्टी" in gav or "सुट्टी" in vasti:
                holiday_row = list(row[10:])
            else:
                data_rows.append((gav, vasti, list(row[10:])))

        if holiday_row:
            h_sets = [parse_holiday_cell(v) for v in holiday_row]
        else:
            h_sets = [set()] * 24

        half_meta = []
        for i, col_label in enumerate(month_cols):
            label     = str(col_label).strip()
            dash      = label.rfind('-')
            mr_name   = label[:dash].strip() if dash != -1 else label
            code      = label[dash+1:].strip() if dash != -1 else ""
            month_num = mr_month_map.get(mr_name, 0)
            part      = 1 if i % 2 == 0 else 2
            half_meta.append({
                "month_num": month_num,
                "part":      part,
                "code":      code,
                "label":     label,
                "col_idx":   i,
                "holidays":  h_sets[i] if i < len(h_sets) else set(),
            })
        return data_rows, half_meta

    # ── Load Holidays.xlsx (cached) ───────────────────────────────────────────
    @st.cache_data
    def load_holidays(path_str):
        """
        Returns dict: { (month, day): "holiday name" }
        Works for any year present in the file.
        """
        if not Path(path_str).exists():
            return {}
        df = pd.read_excel(path_str, header=0)
        df.columns = [c.strip().lower() for c in df.columns]
        result = {}
        for _, row in df.iterrows():
            try:
                dt   = pd.to_datetime(row['date'])
                name = str(row['festive_name']).strip()
                result[(dt.month, dt.day)] = name
            except Exception:
                continue
        return result

    data_rows, half_meta = load_movement(str(excel_path))

    # Build holiday map: prefer Holidays.xlsx; fallback to movement.xlsx
    holidays_from_xlsx = load_holidays(str(holiday_path)) if holiday_path.exists() else {}

    EN_MONTHS = ["January","February","March","April","May","June",
                 "July","August","September","October","November","December"]
    MR_MONTHS = ["जानेवारी","फेब्रुवारी","मार्च","एप्रिल","मे","जून",
                 "जुलै","ऑगस्ट","सप्टेंबर","ऑक्टोबर","नोव्हेंबर","डिसेंबर"]

    # ── Common inputs ─────────────────────────────────────────────────────────
    # Use session_state keys to prevent re-render glitch on widget interaction
    if 'month_en' not in st.session_state:
        st.session_state.month_en = EN_MONTHS[datetime.date.today().month - 1]
    if 'year' not in st.session_state:
        st.session_state.year = datetime.date.today().year

    cols = st.columns([1, 1, 2])
    month_en = cols[0].selectbox(
        "महिना", EN_MONTHS,
        index=EN_MONTHS.index(st.session_state.month_en),
        format_func=lambda x: MR_MONTHS[EN_MONTHS.index(x)],
        key="month_en"
    )
    year = cols[1].number_input(
        "वर्ष", min_value=2000, max_value=2100,
        value=st.session_state.year, key="year"
    )
    sevak_name = cols[2].text_input("आरोग्य सेवक नाव", key="sevak_name")

    c1, c2 = st.columns(2)
    upkendra   = c1.text_input("उपकेंद्राचे नाव",       key="upkendra")
    pra_kendra = c2.text_input("प्राथमिक आरोग्य केंद्र", key="pra_kendra")

    month_num_sel = EN_MONTHS.index(month_en) + 1
    month_name_mr = MR_MONTHS[month_num_sel - 1]

    fixed_gav = data_rows[0][0] if data_rows else "शेळगाव"
    gav_input = st.text_input("गाव (बदलायचे असल्यास बदला)", value=fixed_gav, key="gav_input")

    # ── Build halves for selected month ──────────────────────────────────────
    halves = [m for m in half_meta if m["month_num"] == month_num_sel]
    if len(halves) < 2:
        st.error("Excel मध्ये या महिन्याचा डेटा नाही!")
        return
    half1, half2 = halves[0], halves[1]

    def build_day_vasti_map(col_idx):
        mapping = {}
        for _, vasti, date_vals in data_rows:
            if col_idx < len(date_vals):
                val = date_vals[col_idx]
                if val is not None:
                    try:
                        day = int(val)
                        if day > 0:
                            mapping[day] = vasti
                    except (ValueError, TypeError):
                        pass
        return mapping

    map1 = build_day_vasti_map(half1["col_idx"])
    map2 = build_day_vasti_map(half2["col_idx"])
    day_vasti_map = {**map1, **map2}

    # ── Holiday resolution ────────────────────────────────────────────────────
    # movement.xlsx holidays (day numbers in the month)
    movement_holiday_days = half1["holidays"] | half2["holidays"]

    def is_holiday(day):
        """True if this day is a holiday (from either source)."""
        # From Holidays.xlsx — exact date match for selected month/year
        if (month_num_sel, day) in holidays_from_xlsx:
            return True
        # From movement.xlsx column
        if day in movement_holiday_days:
            return True
        return False

    def get_holiday_name(day):
        """Return holiday name, preferring Holidays.xlsx."""
        key = (month_num_sel, day)
        if key in holidays_from_xlsx:
            return holidays_from_xlsx[key]
        # fallback: search movement.xlsx holidays in floating map
        return "सार्वजनिक सुट्टी"

    # ── Calendar helpers ──────────────────────────────────────────────────────
    total_days = calendar.monthrange(year, month_num_sel)[1]
    dates      = [datetime.date(year, month_num_sel, d) for d in range(1, total_days + 1)]

    tuesdays     = [d for d in dates if d.weekday() == 1]
    mondays      = [d for d in dates if d.weekday() == 0]
    first_monday  = mondays[0]  if mondays  else None
    first_tuesday = tuesdays[0] if tuesdays else None
    third_tuesday = tuesdays[2] if len(tuesdays) >= 3 else None

    # ── Font load ─────────────────────────────────────────────────────────────
    font_path = BASE_DIR / "fonts" / "NotoSerifDevanagari-VariableFont_wdth,wght.ttf"
    if not font_path.exists():
        st.error("❌ Font 'NotoSerifDevanagari-...' missing in fonts/ folder!")
        return
    font_b64 = base64.b64encode(font_path.read_bytes()).decode()

    # ── Row builder (shared logic for Tab1 & Tab2) ────────────────────────────
    def build_row_base(dt):
        """Return dict with all fields; caller adds tab-specific fields."""
        day       = dt.day
        is_sun    = (dt.weekday() == 6)
        is_hol    = is_holiday(day)
        vasti     = day_vasti_map.get(day, "")

        if is_hol:
            return dict(
                दिनांक=dt.strftime("%d-%m-%Y"),
                कोठून="",
                कुठे="सा.सुट्टी",
                कामाचे_स्वरूप=get_holiday_name(day),
                _type="holiday",
            )
        if is_sun:
            return dict(
                दिनांक=dt.strftime("%d-%m-%Y"),
                कोठून="",
                कुठे="---",
                कामाचे_स्वरूप="रविवार",
                _type="sunday",
            )
        if dt == first_monday:
            return dict(
                दिनांक=dt.strftime("%d-%m-%Y"),
                कोठून=gav_input,
                कुठे=vasti,
                कामाचे_स्वरूप="नियमित लसीकरण भुजबळ वस्ती",
                _type="work",
                _firti=True,
            )
        if dt in (first_tuesday, third_tuesday):
            return dict(
                दिनांक=dt.strftime("%d-%m-%Y"),
                कोठून=gav_input,
                कुठे=vasti,
                कामाचे_स्वरूप="नियमित लसीकरण शेळगाव",
                _type="work",
                _firti=True,
            )
        # Regular work day
        kothe  = (f"घरभेट व कंटेनर सर्वेक्षण {vasti}"
                  if vasti and vasti not in ("---", "")
                  else "बहुउद्देशीय कार्य प्रा. आ. केंद्र शेळगाव")
        return dict(
            दिनांक=dt.strftime("%d-%m-%Y"),
            कोठून=gav_input,
            कुठे=vasti,
            कामाचे_स्वरूप=kothe,
            _type="work",
            _firti=bool(vasti and vasti not in ("---", "")),
        )

    # ════════════════════════════════════════════════════════════════════════
    # TABS
    # ════════════════════════════════════════════════════════════════════════
    tab1, tab2 = st.tabs(["📅 आगाऊ फिरती कार्यक्रम", "📒 मासिक दैनंदिनी"])

    # ── TAB 1 ─────────────────────────────────────────────────────────────────
    with tab1:
        st.markdown("### 📅 मासिक आगाऊ फिरती कार्यक्रम")

        rows1 = []
        for dt in dates:
            base = build_row_base(dt)
            rows1.append({
                "दिनांक":         base["दिनांक"],
                "कोठून":          base["कोठून"],
                "कुठे":           base["कुठे"],
                "कामाचे स्वरूप": base["कामाचे_स्वरूप"],
                "शेरा":           "",
            })

        df1 = pd.DataFrame(rows1)
        edited_df1 = st.data_editor(
            df1,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "दिनांक":         st.column_config.TextColumn(width="small",  disabled=True),
                "कोठून":          st.column_config.TextColumn(width="small"),
                "कुठे":           st.column_config.TextColumn(width="medium"),
                "कामाचे स्वरूप": st.column_config.TextColumn(width="large"),
                "शेरा":           st.column_config.TextColumn(width="medium"),
            },
            key="df1_editor",
        )

        if holidays_from_xlsx:
            month_hols = {d: name for (m, d), name in holidays_from_xlsx.items() if m == month_num_sel}
            if month_hols:
                hol_str = ", ".join(f"{d} ({name})" for d, name in sorted(month_hols.items()))
                st.success(f"🗓️ **Holidays.xlsx सुट्ट्या:** {hol_str}")
        st.info(
            f"📋 **Excel कोड:** `{half1['label']}` | `{half2['label']}`\n\n"
            f"🎉 **Movement सुट्ट्या (भाग १):** {sorted(half1['holidays']) or 'नाही'}  \n"
            f"🎉 **Movement सुट्ट्या (भाग २):** {sorted(half2['holidays']) or 'नाही'}"
        )

        # PDF 1
        data1_json = edited_df1.to_dict(orient="records")
        json1_js   = json.dumps(data1_json, ensure_ascii=False)

        components.html(
            f"""
            <html><head><meta charset="utf-8"/>
              <script src="https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.72/pdfmake.min.js"></script>
              <script src="https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.72/vfs_fonts.js"></script>
            </head>
            <body style="font-family:sans-serif;padding:10px;">
              <button onclick="previewPDF1()"
                style="padding:9px 18px;background:#2196F3;color:#fff;border:none;border-radius:6px;margin-right:10px;font-size:14px;cursor:pointer;">
                👁 Preview
              </button>
              <button onclick="downloadPDF1()"
                style="padding:9px 18px;background:#4CAF50;color:#fff;border:none;border-radius:6px;font-size:14px;cursor:pointer;">
                ⬇ Download PDF
              </button>
              <script>
                const data1 = {json1_js};
                pdfMake.vfs["Marathi.ttf"] = "{font_b64}";
                pdfMake.fonts = {{ MarathiFont: {{ normal:"Marathi.ttf", bold:"Marathi.ttf", italics:"Marathi.ttf", bolditalics:"Marathi.ttf" }} }};

                const HDR="#BBDEFB", SUTTI="#FFF9C4", RAVI="#FCE4EC";

                function buildDoc1() {{
                  const body = [[
                    {{ text:"दिनांक",                        bold:true, alignment:"center", fillColor:HDR }},
                    {{ text:"कोठून",                         bold:true, alignment:"center", fillColor:HDR }},
                    {{ text:"कोठे",                          bold:true, alignment:"center", fillColor:HDR }},
                    {{ text:"करावयाच्या कामाचा तपशील",      bold:true, alignment:"center", fillColor:HDR }},
                    {{ text:"शेरा",                          bold:true, alignment:"center", fillColor:HDR }},
                  ]];
                  data1.forEach(r => {{
                    const vasti = r["कुठे"] || "";
                    const kothe = r["कामाचे स्वरूप"] || "";
                    const fc    = vasti==="सा.सुट्टी" ? SUTTI : kothe==="रविवार" ? RAVI : null;
                    body.push([
                      {{ text:r["दिनांक"]||"",         alignment:"center", fillColor:fc }},
                      {{ text:r["कोठून"]||"",          alignment:"center", fillColor:fc }},
                      {{ text:vasti,                    alignment:"center", fillColor:fc }},
                      {{ text:kothe,                    alignment:"left",   fillColor:fc }},
                      {{ text:r["शेरा"]||"",           alignment:"center", fillColor:fc }},
                    ]);
                  }});
                  return {{
                    pageSize:"A4", pageMargins:[30,20,30,30],
                    defaultStyle:{{ font:"MarathiFont", fontSize:10 }},
                    content:[
                      {{ text:"प्रा. आ. केंद्र {pra_kendra}                       उपकेंद्र {upkendra}", fontSize:13, alignment:"center", margin:[0,0,0,2] }},
                      {{ text:"मासिक आगाऊ फिरती कार्यक्रम", fontSize:13, bold:true, alignment:"center", margin:[0,0,0,2] }},
                      {{ text:"कर्मचारी नाव: {sevak_name}                                                                     {month_name_mr} {year}", alignment:"left", fontSize:11, bold:true, margin:[0,0,0,6] }},
                      {{ table:{{ widths:["13%","13%","17%","40%","17%"], body }}, layout:{{ hLineWidth:()=>0.6, vLineWidth:()=>0.6, paddingTop:()=>2, paddingBottom:()=>2, paddingLeft:()=>3, paddingRight:()=>3 }} }},
                      {{ columns:[
                          {{ text:"\\n\\n आरोग्य सेवक: {sevak_name}\\n    उपकेंद्र: {upkendra}", fontSize:10, width:"33%", alignment:"center" }},
                          {{ text:"\\n\\n मा. आरोग्य सहाय्यक\\nप्रा. आ. केंद्र: {pra_kendra}",    width:"33%", alignment:"center", fontSize:10, alignment:"center"  }},
                          {{ text:"\\n\\n मा. वैद्यकीय अधिकारी\\nप्रा. आ. केंद्र: {pra_kendra}", fontSize:10, width:"33%", alignment:"center" }},
                        ], margin:[0,14,0,0]
                      }},
                    ],
                  }};
                }}
                function previewPDF1()  {{ pdfMake.createPdf(buildDoc1()).open(); }}
                function downloadPDF1() {{ pdfMake.createPdf(buildDoc1()).download("Agau_Firti_{month_name_mr}_{year}.pdf"); }}
              </script>
            </body></html>
            """,
            height=80,
        )

    # ── TAB 2 ─────────────────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📒 मासिक दैनंदिनी")

        rows2      = []
        work_days  = 0
        firti_days = 0
        sutti_days = 0

        for dt in dates:
            base = build_row_base(dt)
            t    = base.get("_type", "work")

            if t in ("holiday", "sunday"):
                sutti_days += 1
                rows2.append({
                    "दिनांक":         base["दिनांक"],
                    "कोठून":          base["कोठून"],
                    "कुठे":           base["कुठे"],
                    "कामाचे स्वरूप": base["कामाचे_स्वरूप"],
                    "निघण्याची वेळ":  "",
                    "परतीची वेळ":     "",
                    "शेरा":           "",
                })
            else:
                work_days += 1
                if base.get("_firti"):
                    firti_days += 1
                rows2.append({
                    "दिनांक":         base["दिनांक"],
                    "कोठून":          base["कोठून"],
                    "कुठे":           base["कुठे"],
                    "कामाचे स्वरूप": base["कामाचे_स्वरूप"],
                    "निघण्याची वेळ":  "9:00",
                    "परतीची वेळ":     "5:00",
                    "शेरा":           "",
                })

        df2 = pd.DataFrame(rows2)
        edited_df2 = st.data_editor(
            df2,
            num_rows="dynamic",
            use_container_width=True,
            column_config={
                "दिनांक":         st.column_config.TextColumn(width="small",  disabled=True),
                "कोठून":          st.column_config.TextColumn(width="small"),
                "कुठे":           st.column_config.TextColumn(width="medium"),
                "कामाचे स्वरूप": st.column_config.TextColumn(width="large"),
                "निघण्याची वेळ":  st.column_config.TextColumn(width="small"),
                "परतीची वेळ":     st.column_config.TextColumn(width="small"),
                "शेरा":           st.column_config.TextColumn(width="medium"),
            },
            key="df2_editor",
        )

        # Summary stats
        st.markdown("---")
        st.markdown("#### 📊 महिन्याचा सारांश")
        sc1, sc2, sc3, sc4 = st.columns(4)
        ek_kam   = sc1.number_input("एकूण कामाचे दिवस",   value=work_days,  min_value=0, key="ek_kam")
        ek_firti = sc2.number_input("एकूण फिरतीचे दिवस",  value=firti_days, min_value=0, key="ek_firti")
        ek_sutti = sc3.number_input("एकूण सुट्टीचे दिवस", value=sutti_days, min_value=0, key="ek_sutti")
        ek_raje  = sc4.number_input("एकूण रजेचे दिवस",    value=0,          min_value=0, key="ek_raje")

        # PDF 2
        data2_json = edited_df2.to_dict(orient="records")
        json2_js   = json.dumps(data2_json, ensure_ascii=False)

        components.html(
            f"""
            <html><head><meta charset="utf-8"/>
              <script src="https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.72/pdfmake.min.js"></script>
              <script src="https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.72/vfs_fonts.js"></script>
            </head>
            <body style="font-family:sans-serif;padding:10px;">
              <button onclick="previewPDF2()"
                style="padding:9px 18px;background:#2196F3;color:#fff;border:none;border-radius:6px;margin-right:10px;font-size:14px;cursor:pointer;">
                👁 Preview
              </button>
              <button onclick="downloadPDF2()"
                style="padding:9px 18px;background:#4CAF50;color:#fff;border:none;border-radius:6px;font-size:14px;cursor:pointer;">
                ⬇ Download PDF
              </button>
              <script>
                const data2 = {json2_js};
                pdfMake.vfs["Marathi.ttf"] = "{font_b64}";
                pdfMake.fonts = {{ MarathiFont: {{ normal:"Marathi.ttf", bold:"Marathi.ttf", italics:"Marathi.ttf", bolditalics:"Marathi.ttf" }} }};

                const HDR2="#BBDEFB", SUTTI2="#FFF9C4", RAVI2="#FCE4EC";

                function buildDoc2() {{
                  const body2 = [[
                    {{ text:"दिनांक",                  bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"कोठून",                   bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"कोठे",                    bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"निघण्याची वेळ",           bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"परतीची वेळ",              bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"केलेल्या कामाचा तपशील",  bold:true, alignment:"center", fillColor:HDR2 }},
                    {{ text:"शेरा",                    bold:true, alignment:"center", fillColor:HDR2 }},
                  ]];
                  data2.forEach(r => {{
                    const vasti = r["कुठे"] || "";
                    const kothe = r["कामाचे स्वरूप"] || "";
                    const fc    = vasti==="सा.सुट्टी" ? SUTTI2 : kothe==="रविवार" ? RAVI2 : null;
                    body2.push([
                      {{ text:r["दिनांक"]||"",         alignment:"center", fillColor:fc }},
                      {{ text:r["कोठून"]||"",          alignment:"center", fillColor:fc }},
                      {{ text:vasti,                    alignment:"center", fillColor:fc }},
                      {{ text:r["निघण्याची वेळ"]||"",  alignment:"center", fillColor:fc }},
                      {{ text:r["परतीची वेळ"]||"",     alignment:"center", fillColor:fc }},
                      {{ text:kothe,                    alignment:"left",   fillColor:fc }},
                      {{ text:r["शेरा"]||"",           alignment:"center", fillColor:fc }},
                    ]);
                  }});

                  const statsBody = [[
                    {{ text:"एकूण कामाचे दिवस",   bold:true, alignment:"center", fillColor:"#E3F2FD" }},
                    {{ text:"एकूण फिरतीचे दिवस",  bold:true, alignment:"center", fillColor:"#E3F2FD" }},
                    {{ text:"एकूण सुट्टीचे दिवस", bold:true, alignment:"center", fillColor:"#E3F2FD" }},
                    {{ text:"एकूण रजेचे दिवस",    bold:true, alignment:"center", fillColor:"#E3F2FD" }},
                  ],[
                    {{ text:"{ek_kam}",   alignment:"center", fontSize:11, bold:true }},
                    {{ text:"{ek_firti}", alignment:"center", fontSize:11, bold:true }},
                    {{ text:"{ek_sutti}", alignment:"center", fontSize:11, bold:true }},
                    {{ text:"{ek_raje}",  alignment:"center", fontSize:11, bold:true }},
                  ]];

                  const sigBody = [[
                    {{ text:"आरोग्य सेवक\\n{sevak_name}\\nउपकेंद्र: {upkendra}",      alignment:"center", fontSize:9 }},
                    {{ text:"मा. आरोग्य सहाय्यक\\nप्रा. आ. केंद्र: {pra_kendra}",     alignment:"center", fontSize:9 }},
                    {{ text:"मा. वैद्यकीय अधिकारी\\nप्रा. आ. केंद्र: {pra_kendra}",   alignment:"center", fontSize:9 }},
                  ]];

                  return {{
                    pageSize:"A4", pageMargins:[25,20,25,5],
                    defaultStyle:{{ font:"MarathiFont", fontSize:9 }},
                    content:[
                      {{ text:"प्रा. आ. केंद्र {pra_kendra}                       उपकेंद्र {upkendra}", fontSize:13, alignment:"center", margin:[0,0,0,2] }},
                      {{ text:"मासिक दैनंदिनी", fontSize:14, bold:true, alignment:"center", margin:[0,0,0,2] }},
                      {{ text:"कर्मचारी नाव: {sevak_name}                                                                          {month_name_mr} {year}", alignment:"left", fontSize:11, bold:true, margin:[0,0,0,6] }},
                      {{ table:{{ widths:["11%","11%","17%","10%","10%","31%","11%"], body:body2 }}, layout:{{ hLineWidth:()=>0.6, vLineWidth:()=>0.6, paddingTop:()=>2, paddingBottom:()=>2, paddingLeft:()=>3, paddingRight:()=>3 }} }},
                      {{ text:"", margin:[0,4,0,4] }},
                      {{ table:{{ widths:["25%","25%","25%","25%"], body:statsBody }}, layout:{{ hLineWidth:()=>0.6, vLineWidth:()=>0.6, paddingTop:()=>2, paddingBottom:()=>2 }} }},
                      {{ text:"", margin:[0,16,0,25] }},
                      {{ table:{{ widths:["33%","33%","34%"], body:sigBody }}, layout:"noBorders" }},
                    ],
                  }};
                }}
                function previewPDF2()  {{ pdfMake.createPdf(buildDoc2()).open(); }}
                function downloadPDF2() {{ pdfMake.createPdf(buildDoc2()).download("Masik_Daindini_{month_name_mr}_{year}.pdf"); }}
              </script>
            </body></html>
            """,
            height=80,
        )


if __name__ == "__main__":
    import streamlit as st
    st.set_page_config(page_title="मासिक डायरी", layout="wide")
    monthly_diary()